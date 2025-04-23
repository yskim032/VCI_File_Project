import sys
import os
import pandas as pd
from collections import defaultdict
from typing import List, Dict, Set, Tuple
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                           QHBoxLayout, QLabel, QTextEdit, QPushButton, 
                           QFileDialog, QMessageBox, QTabWidget,
                           QFrame, QRadioButton, QButtonGroup)
from PyQt5.QtCore import Qt, QMimeData
from PyQt5.QtGui import QDragEnterEvent, QDropEvent
from openpyxl.styles import PatternFill
from openpyxl import Workbook

class ContainerAnalyzer:
    def __init__(self, operation_type: str, tpf_containers: Set[str], truck_containers: Set[str]):
        """
        Initialize ContainerAnalyzer
        
        Args:
            operation_type: 'DIS' or 'LOD'
            tpf_containers: Set of container numbers for TPF
            truck_containers: Set of container numbers for Truck
        """
        if operation_type not in ['DIS', 'LOD']:
            raise ValueError("operation_type must be either 'DIS' or 'LOD'")
            
        self.operation_type = operation_type
        self.tpf_containers = tpf_containers
        self.truck_containers = truck_containers
        self.container_groups = defaultdict(list)

    def _extract_container_info(self, line: str) -> Tuple[str, str, str, str]:
        """Extract container information from ASC file line"""
        return (
            line[6:17].strip(),   # container_number (MSBU3091780)
            line[44:48].strip(),  # container_type (45-48 position)
            line[51:52].strip(),  # full_empty (52 position)
            line[19:22].strip(),  # operator_code (MSC)
        )

    def parse_container_data(self, line: str) -> Dict:
        """Parse a single line from ASC file"""
        container_number, container_type, full_empty, operator_code = self._extract_container_info(line)

        # Extract weight from line[48:51] and convert to float
        raw_weight = line[48:51].strip()
        try:
            weight = float(raw_weight) / 10
        except ValueError:
            weight = 0.0

        # Check for IMO container (line[60:64])
        is_imo = 'Yes' if line[60:64].strip() else 'NO'

        # Check for OOG container in specified ranges
        oog_ranges = [
            line[92:95].strip(),  # [92,3]
            line[95:98].strip(),  # [95,3]
            line[98:101].strip(), # [98,3]
            line[101:104].strip(), # [101,3]
            line[104:107].strip()  # [104,3]
        ]
        is_oog = 'Yes' if any(oog_ranges) else 'No'

        # Only print first 50 containers
        if sum(len(containers) for containers in self.container_groups.values()) < 50:
            print(f"Container: {container_number}, weight: {weight}")
       
        # Create container key based on all fields that affect grouping
        group_key = (
            self.operation_type,
            container_type,
            full_empty,
            operator_code,
            is_oog,  # OOG
            'No',  # Damaged
            'No',  # SOC
            'No',  # Coastal Cargo
            'No',  # To Rail
            'No',  # To Barge
            'Yes' if container_number in self.tpf_containers else 'No',  # To TPF
            'Yes' if container_number in self.truck_containers else 'No',  # To Truck
            'No',   # Not for MSC Account
            is_imo  # IMO
        )

        return {
            'container_number': container_number,
            'group_key': group_key,
            'weight': weight
        }

    def process_file(self, file_path: str) -> pd.DataFrame:
        """
        Process ASC file and return summary DataFrame
        
        Args:
            file_path: Path to ASC file
            
        Returns:
            DataFrame with container summary
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    if not line.startswith('$'):  # Skip header lines
                        container_data = self.parse_container_data(line)
                        # Only include containers with MSC operator code
                        if container_data['group_key'][3] == 'MSC':
                            self.container_groups[container_data['group_key']].append(
                                (container_data['container_number'], container_data['weight'])
                            )
        except FileNotFoundError:
            raise FileNotFoundError(f"ASC file not found: {file_path}")
        except Exception as e:
            raise Exception(f"Error processing ASC file: {str(e)}")

        # Create summary records
        summary_records = []
        for group_key, containers in self.container_groups.items():
            # Calculate total weight for the group and round to nearest integer
            total_weight = round(sum(weight for _, weight in containers))
            
            record = {
                'Operation': group_key[0],
                'Container Type': group_key[1],
                'Full/Empty': group_key[2],
                'Operator Code': group_key[3],
                'Weight': total_weight,
                'Quantity': len(containers),
                'OOG': group_key[4],
                'Damaged': group_key[5],
                'IMO': group_key[13],
                'SOC': group_key[6],
                'Coastal Cargo': group_key[7],
                'To Rail': group_key[8],
                'To Barge': group_key[9],
                'To TPF': group_key[10],
                'To Truck': group_key[11],
                'Not for MSC Account': group_key[12]
            }
            summary_records.append(record)

        # Create DataFrame
        df = pd.DataFrame(summary_records)
        column_order = [
            'Operation', 'Container Type', 'Full/Empty', 'Operator Code', 'Weight',
            'Quantity', 'OOG', 'Damaged', 'IMO', 'SOC', 'Coastal Cargo', 'To Rail',
            'To Barge', 'To TPF', 'To Truck', 'Not for MSC Account'
        ]
        return df[column_order]

def create_summary(asc_file: str, operation_type: str, 
                  tpf_containers: List[str], truck_containers: List[str], 
                  output_file: str = None) -> None:
    """
    Create container summary Excel file
    
    Args:
        asc_file: Path to ASC file
        operation_type: 'DIS' or 'LOD'
        tpf_containers: List of container numbers for TPF
        truck_containers: List of container numbers for Truck
        output_file: Path to output Excel file (optional, defaults to ASC filename with .xlsx extension)
    """
    try:
        # Convert container lists to sets for faster lookup
        tpf_set = set(tpf_containers)
        truck_set = set(truck_containers)

        # Create analyzer and process file
        analyzer = ContainerAnalyzer(operation_type, tpf_set, truck_set)
        summary_df = analyzer.process_file(asc_file)

        # Use dragged ASC filename for output if not specified
        if output_file is None:
            # Get just the filename from the full path
            asc_filename = os.path.basename(asc_file)
            output_file = asc_filename.replace('.ASC', '.xlsx')

        # Create Excel writer with openpyxl engine
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write DataFrame to Excel
            summary_df.to_excel(writer, index=False, sheet_name='Summary')
            
            # Get the worksheet
            worksheet = writer.sheets['Summary']
            
            # Define pink fill
            pink_fill = PatternFill(start_color='FFFFC0CB', end_color='FFFFC0CB', fill_type='solid')
            
            # Apply pink fill to cells containing 'Yes'
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    if cell.value == 'Yes':
                        cell.fill = pink_fill

        print(f"Summary successfully written to {output_file}")
        
    except Exception as e:
        print(f"Error creating summary: {str(e)}")
        raise

class DropArea(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setFrameStyle(QFrame.StyledPanel | QFrame.Sunken)
        self.setStyleSheet("""
            QFrame {
                background-color: #f8f9fa;
                border: 2px dashed #dee2e6;
                border-radius: 5px;
                padding: 20px;
                min-height: 100px;
            }
        """)
        
        layout = QVBoxLayout(self)
        self.label = QLabel("여기에 ASC 파일을 드래그 앤 드롭하세요")
        self.label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label)
        
        self.file_path = None
        
    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            
    def dropEvent(self, event: QDropEvent):
        files = [url.toLocalFile() for url in event.mimeData().urls()]
        if files:
            self.file_path = files[0]
            self.label.setText(os.path.basename(self.file_path))
            self.label.setStyleSheet("color: #28a745;")  # Green color for success

class ContainerTab(QWidget):
    def __init__(self, title, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(10)
        
        label = QLabel(f"{title} (한 줄에 하나의 컨테이너 번호):")
        self.text_edit = QTextEdit()
        self.text_edit.setMinimumHeight(200)
        self.text_edit.setStyleSheet("""
            QTextEdit {
                background-color: white;
                border: 1px solid #ced4da;
                border-radius: 4px;
                padding: 5px;
            }
        """)
        
        layout.addWidget(label)
        layout.addWidget(self.text_edit)
        layout.addStretch()
        
    def get_container_list(self):
        return [line.strip() for line in self.text_edit.toPlainText().split('\n') if line.strip()]

class ContainerAnalyzerGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('Container Analyzer')
        self.setGeometry(100, 100, 1000, 800)
        
        # Create central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(20)
        
        # Operation type selection with radio buttons
        op_layout = QHBoxLayout()
        op_label = QLabel('Operation Type:')
        op_layout.addWidget(op_label)
        
        self.op_group = QButtonGroup(self)
        self.discharge_radio = QRadioButton('Discharge')
        self.load_radio = QRadioButton('Load')
        self.discharge_radio.setChecked(True)  # Default selection
        
        self.op_group.addButton(self.discharge_radio)
        self.op_group.addButton(self.load_radio)
        
        op_layout.addWidget(self.discharge_radio)
        op_layout.addWidget(self.load_radio)
        op_layout.addStretch()
        
        main_layout.addLayout(op_layout)
        
        # ASC file drop area
        main_layout.addWidget(QLabel('ASC File:'))
        self.drop_area = DropArea()
        main_layout.addWidget(self.drop_area)
        
        # Tab widget
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #dee2e6;
                border-radius: 4px;
                background: white;
            }
            QTabBar::tab {
                padding: 8px 16px;
                margin: 2px;
                background: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
            }
            QTabBar::tab:selected {
                background: white;
                border-bottom: none;
            }
        """)
        
        # Create tabs
        self.tpf_tab = ContainerTab('TPF Containers')
        self.truck_tab = ContainerTab('Truck Containers')
        self.local_tab = ContainerTab('Local')
        self.same_ts_tab = ContainerTab('Same TS')
        self.external_ts_tab = ContainerTab('External TS')
        
        self.tab_widget.addTab(self.tpf_tab, 'TPF Containers')
        self.tab_widget.addTab(self.truck_tab, 'Truck Containers')
        self.tab_widget.addTab(self.local_tab, 'Local')
        self.tab_widget.addTab(self.same_ts_tab, 'Same TS')
        self.tab_widget.addTab(self.external_ts_tab, 'External TS')
        
        main_layout.addWidget(self.tab_widget)
        
        # Process button
        self.process_btn = QPushButton('Create Summary')
        self.process_btn.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                padding: 10px 20px;
                font-size: 14px;
                border: none;
                border-radius: 5px;
                min-height: 40px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """)
        self.process_btn.clicked.connect(self.process_data)
        main_layout.addWidget(self.process_btn)

    def get_operation_type(self, tab_name: str) -> tuple:
        """Get operation type and truck flag based on radio selection and tab"""
        is_discharge = self.discharge_radio.isChecked()
        
        if tab_name == 'Local':
            return ('DIS' if is_discharge else 'LOD', False)
        elif tab_name == 'Same TS':
            return ('TSD' if is_discharge else 'TSL', False)
        elif tab_name == 'External TS':
            return ('TSD' if is_discharge else 'TSL', True)
        return (None, False)
        
    def process_data(self):
        try:
            # Get ASC file path
            if not self.drop_area.file_path:
                QMessageBox.warning(self, 'Error', 'ASC 파일을 드래그 앤 드롭 해주세요.')
                return
                
            asc_file = self.drop_area.file_path
            if not os.path.exists(asc_file):
                QMessageBox.warning(self, 'Error', 'ASC 파일을 찾을 수 없습니다.')
                return

            # Get container lists from tabs
            tpf_containers = self.tpf_tab.get_container_list()
            truck_containers = self.truck_tab.get_container_list()
            
            # Add containers from operation tabs based on rules
            local_containers = self.local_tab.get_container_list()
            same_ts_containers = self.same_ts_tab.get_container_list()
            external_ts_containers = self.external_ts_tab.get_container_list()
            
            # Add External TS containers to truck containers
            truck_containers.extend(external_ts_containers)
            
            # Get operation type based on radio selection
            operation_type = 'DIS' if self.discharge_radio.isChecked() else 'LOD'
            
            # Create output file path using ASC filename
            asc_filename = os.path.basename(asc_file)
            output_file = asc_filename.replace('.ASC', '.xlsx')
            output_path = os.path.join(os.path.dirname(asc_file), output_file)
            
            # Create summary
            create_summary(
                asc_file=asc_file,
                operation_type=operation_type,
                tpf_containers=tpf_containers,
                truck_containers=truck_containers,
                output_file=output_path
            )
            
            QMessageBox.information(
                self, 
                'Success', 
                f'Summary가 성공적으로 생성되었습니다:\n{output_path}'
            )
                
        except Exception as e:
            QMessageBox.critical(self, 'Error', str(e))

if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Fusion')  # Modern style
    window = ContainerAnalyzerGUI()
    window.show()
    sys.exit(app.exec_()) 